from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook, load_workbook
from datetime import datetime
from time import sleep
import schedule
import os

def iniciar_driver():
    print("Driver do navegador iniciado.")
    chrome_options = Options()
    arguments = ['--lang=pt-BR', '--window-size=800,600', '--incognito']
    for argument in arguments:
        chrome_options.add_argument(argument)
        
    chrome_options.add_experimental_option('prefs', {
        'download.prompt_for_download': False,
        'profile.default_content_setting_values.notifications': 2,
        'profile.default_content_setting_values.automatic_downloads': 1,
    })

    driver = webdriver.Chrome(options=chrome_options)
    return driver

def site_produto(driver, site):
    print("Acessando o site do produto...")
    driver.get(site)
    sleep(10)
    
    try:
        element_precos = driver.find_elements(By.CSS_SELECTOR, 'span.andes-money-amount__fraction')
        if len(element_precos) >= 2:
            preco = element_precos[1].text
            print(f"Preço capturado com sucesso: R${preco}")
            return preco
        else:
            return None
    except NoSuchElementException:
        return None

def atualizar_planilha(preco, site):
    print("Atualizando a planilha com as novas informações...")
    arquivo_planilha = 'desafio_destrava_dev3.xlsx'
    if os.path.exists(arquivo_planilha):
        print("Planilha existente encontrada. Carregando...")
        workbook = load_workbook(arquivo_planilha)
        sheet_produto = workbook['produto']
    else:
        workbook = Workbook()
        sheet_produto = workbook.active
        sheet_produto.title = 'produto'
        sheet_produto.append(['produto', 'data', 'valor', 'link'])
    
    data_atual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet_produto.append(['iPhone 15 Pro Max', data_atual, preco, site])
    workbook.save(arquivo_planilha)
    print("Planilha atualizada e salva com sucesso.")

def main():
    print("Iniciando a execução do script principal...")
    driver = iniciar_driver()
    site = 'https://www.mercadolivre.com.br/apple-iphone-15-pro-max-256-gb-titnio-preto-distribuidor-autorizado/p/MLB1027172724#searchVariation%3DMLB1027172724%26position%3D2%26search_layout%3Dstack%26type%3Dproduct%26tracking_id%3D797ac78e-6fc1-4ee3-86f2-b1aa7d2c6f7c'
    preco = site_produto(driver, site)
    if preco is not None:
        atualizar_planilha(preco, site)
    driver.quit()
    print("Execução do script principal concluída.")
    
if __name__ == "__main__":
    print("Executando pela primeira vez...")
    main()

    print("Agendando a execução a cada 30 minutos...")
    schedule.every(30).minutes.do(main)

    while True:
        schedule.run_pending()
        sleep(1)
